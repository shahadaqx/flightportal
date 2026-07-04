import streamlit as st
import pandas as pd
from datetime import datetime, time
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.title("✈️ Portal Data Formatter (with Template Paste)")

# Final smart rollover logic
def format_datetime(date, raw_time, base_time=None):
    if pd.isna(date) or pd.isna(raw_time):
        return None
    try:
        base_date = pd.to_datetime(date).date()

        def to_time(val):
            if isinstance(val, str):
                return datetime.strptime(val.strip(), "%H:%M").time()
            elif isinstance(val, time):
                return val
            else:
                return None

        raw_time_obj = to_time(raw_time)
        base_time_obj = to_time(base_time)

        if raw_time_obj is None:
            return None

        if base_time_obj and base_time_obj >= time(18, 0) and raw_time_obj < time(3, 0):
            base_date += pd.Timedelta(days=1)

        full_datetime = datetime.combine(base_date, raw_time_obj.replace(second=0))
        return full_datetime
    except Exception:
        return None


def get_customer(x):
    """Safely derive the customer code from the flight number.
    Returns None instead of crashing when the value is missing/blank."""
    if not isinstance(x, str):
        return None
    x = x.strip()
    if x == '' or x.lower() == 'nan':
        return None
    return 'XLR' if x.startswith('DHX') else x[:2]


def extract_services(row):
    services = []
    for col in row.index:
        if isinstance(col, str) and str(row[col]).strip() == '√':
            services.append(col.strip())

    remark = str(row.get('OTHER SERVICES/REMARKS', '')).upper()
    if 'ON CALL - NEEDED ENGINEER SUPPORT' in remark:
        services.append('On Call')
    elif 'CANCELED' in remark or 'CANCELLED' in remark:
        # Any cancellation (with or without info) is simply labeled "Cancelled Flight"
        services.append('Cancelled Flight')
    elif 'ON CALL' in remark:
        services.append('Per Landing')

    corrected_services = []
    for service in services:
        if service == 'TECH. SUPT':
            corrected_services.append('TECH SUPPORT')
        elif service == 'HEAD SET':
            corrected_services.append('Headset')
        else:
            corrected_services.append(service)

    return ', '.join(corrected_services) if corrected_services else None


def categorize(row):
    remark = str(row.get('OTHER SERVICES/REMARKS', '')).upper()
    if 'TRANSIT' in remark:
        return '1_TRANSIT'
    elif 'ON CALL - NEEDED ENGINEER SUPPORT' in remark:
        return '2_ONCALL_ENGINEER'
    elif ('CANCELED WITHOUT INFORMATION' in remark or 'CANCELLED WITHOUT INFORMATION' in remark
          or 'CANCELED WITH INFORMATION' in remark or 'CANCELLED WITH INFORMATION' in remark
          or 'CANCELED' in remark or 'CANCELLED' in remark):
        return '3_CANCELED'
    elif 'ON CALL' in remark:
        return '4_ONCALL_RECORDED'
    else:
        return '5_OTHER'


def blank_to_none(x):
    """Treat whitespace-only cells (stray spaces left in the sheet) as truly empty."""
    if isinstance(x, str) and x.strip() == '':
        return None
    return x


def find_report_sheet(uploaded_file, expected_name='Daily Operations Report'):
    """Find the correct sheet to read, tolerating case/whitespace differences
    in the sheet name (e.g. 'daily operations report ', 'DAILY OPERATIONS REPORT').
    Falls back to the first sheet if no reasonable match is found, and warns
    the user either way so this never fails silently."""
    xl = pd.ExcelFile(uploaded_file)
    sheet_names = xl.sheet_names

    for name in sheet_names:
        if name.strip().lower() == expected_name.strip().lower():
            return name, sheet_names

    # No exact (case/whitespace-insensitive) match - try a loose "contains" match
    for name in sheet_names:
        if 'daily operations' in name.strip().lower():
            st.warning(f"⚠️ Sheet named exactly '{expected_name}' not found. "
                       f"Using closest match: '{name}'.")
            return name, sheet_names

    # Still nothing - fall back to the first sheet so the app doesn't crash outright
    st.warning(f"⚠️ Could not find a sheet named '{expected_name}' in the uploaded file. "
               f"Available sheets: {sheet_names}. Falling back to the first sheet: '{sheet_names[0]}'.")
    return sheet_names[0], sheet_names


def process_file(uploaded_file, template_file):
    sheet_name, available_sheets = find_report_sheet(uploaded_file)
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=4)

    # Normalize stray whitespace-only cells to real blanks BEFORE checking for empty rows,
    # so leftover artifact rows (e.g. a single stray space in one cell) get dropped correctly.
    df = df.map(blank_to_none)
    df.dropna(how='all', inplace=True)

    # Extra safety net: if a row has neither a DATE nor a FLT NO., it isn't a real flight
    # record (confirmed real cancellations always carry both) - drop it so it can't crash
    # downstream processing.
    df.dropna(subset=['DATE', 'FLT NO.'], how='all', inplace=True)

    df.rename(columns=lambda x: x.strip() if isinstance(x, str) else x, inplace=True)
    df.rename(columns={
        'REG.': 'REG',
        'TECH.\nSUPT': 'TECH. SUPT',
        'TECH. SUPT': 'TECH SUPPORT',
        'HEAD SET': 'Headset',
        'TRANSIT': 'Transit',
        'WKLY CK': 'Weekly Check',
        'DAILY CK': 'Daily Check'
    }, inplace=True)

    df['STA.'] = df.apply(lambda row: format_datetime(row['DATE'], row['STA'], None), axis=1)
    df['ATA.'] = df.apply(lambda row: format_datetime(row['DATE'], row['ATA'], row['STA']), axis=1)
    df['STD.'] = df.apply(lambda row: format_datetime(row['DATE'], row.get('STD'), row['STA']), axis=1)
    df['ATD.'] = df.apply(lambda row: format_datetime(row['DATE'], row.get('ATD'), row['STA']), axis=1)

    canceled_mask = df['OTHER SERVICES/REMARKS'].astype(str).str.contains('CANCELED|CANCELLED', case=False, na=False)
    df.loc[canceled_mask, 'ATA.'] = df.loc[canceled_mask, 'STA.']
    df.loc[canceled_mask, 'ATD.'] = df.loc[canceled_mask, 'STD.']

    df['Customer'] = df['FLT NO.'].astype(str).str.strip().apply(get_customer)
    df['Services'] = df.apply(extract_services, axis=1)
    df['Is Canceled'] = df['OTHER SERVICES/REMARKS'].astype(str).str.contains('CANCELED|CANCELLED', na=False, case=False)
    df['Category'] = df.apply(categorize, axis=1)
    df.sort_values(by=['Category', 'STA.'], inplace=True)

    result_rows = []
    for _, row in df.iterrows():
        try:
            result_rows.append({
                'WO#': row['W/O'],
                'Station': 'KKIA',
                'Customer': row['Customer'],
                'Flight No.': row['FLT NO.'],
                'Registration Code': row['REG'],
                'Aircraft': row['A/C TYPES'],
                'Date': pd.to_datetime(row['DATE']),
                'STA.': row['STA.'],
                'ATA.': row['ATA.'],
                'STD.': row['STD.'],
                'ATD.': row['ATD.'],
                'Is Canceled': row['Is Canceled'],
                'Services': row['Services'],
                'Employees': ', '.join(filter(None, [
                    str(int(row['ENGR'])) if pd.notna(row['ENGR']) and str(row['ENGR']).replace('.', '', 1).isdigit() else '',
                    str(int(row['TECH'])) if pd.notna(row['TECH']) and str(row['TECH']).replace('.', '', 1).isdigit() else ''
                ])),
                'Remarks': '',
                'Comments': ''
            })
        except Exception:
            pass

    result_df = pd.DataFrame(result_rows)

    if result_df.empty:
        st.warning("⚠️ No flight rows were extracted from the uploaded report. "
                    "The output file will be blank - check that the report matches "
                    "the expected column layout (header on row 5).")

    output = io.BytesIO()
    template_wb = load_workbook(template_file)

    # IMPORTANT: target the 'Template' sheet explicitly by name. Using
    # template_wb.active is unreliable - it points to whichever tab was
    # open/selected when the file was last saved (e.g. it could silently
    # be 'Lookups' instead of 'Template'), which would make the actual
    # Template sheet appear completely blank.
    if 'Template' in template_wb.sheetnames:
        ws = template_wb['Template']
    else:
        ws = template_wb.active

    # Write starting after the header row (assumed row 2)
    start_row = 2
    for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value

            # Match the EXACT format confirmed to be accepted by the portal
            # (verified against a manually-filled file the portal approved):
            #   - Date column: real datetime, format 'mm-dd-yy'
            #   - STA./ATA./STD./ATD.: real datetime, format 'm/d/yy h:mm'
            if c_idx == 7 and isinstance(value, (datetime, pd.Timestamp)):
                cell.value = value if isinstance(value, datetime) else value.to_pydatetime()
                cell.number_format = 'mm-dd-yy'
            elif isinstance(value, (datetime, pd.Timestamp)):
                cell.value = value if isinstance(value, datetime) else value.to_pydatetime()
                cell.number_format = 'm/d/yy h:mm'
            elif c_idx in (13, 14):
                # Services / Employees columns: force text format to match accepted file
                cell.number_format = '@'

    template_wb.save(output)
    output.seek(0)

    report_date = df['DATE'].iloc[0] if not df.empty else None
    return output, report_date


# Upload files
uploaded_file = st.file_uploader("Upload Daily Operations Report", type=["xlsx"])
template_file = st.file_uploader("Upload Work Order Template", type=["xlsx"])

if uploaded_file and template_file:
    st.success("✅ Files uploaded successfully!")
    final_output, report_date = process_file(uploaded_file, template_file)

    if report_date is not None:
        try:
            date_obj = pd.to_datetime(report_date)
            filename = date_obj.strftime("%d%b%Y").upper() + "_WorkOrders.xlsx"
        except Exception:
            filename = "Final_WorkOrders.xlsx"
    else:
        filename = "Final_WorkOrders.xlsx"

    st.download_button("📥 Download Final Work Order File", data=final_output, file_name=filename)


    

    

